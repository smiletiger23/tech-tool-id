import os
import openpyxl
from db_manager import FixtureDBManager
import shutil


class ExcelClassifierImporter:
    def __init__(self, db_manager_instance):
        self.db_manager = db_manager_instance
        self.sheet_configs = {
            "Категории": {
                "handler": self.db_manager.add_category,
                "get_all_from_db": self.db_manager.get_categories,
                "columns": ["CategoryCode", "CategoryName"],
                "required_cols": ["CategoryCode", "CategoryName"],
                "key_cols": ["CategoryCode"]
            },
            "Серии": {
                "handler": self.db_manager.add_series_description,
                "get_all_from_db": self.db_manager.get_series_descriptions,
                "columns": ["CategoryCode", "SeriesCode", "SeriesName"],
                "required_cols": ["CategoryCode", "SeriesCode", "SeriesName"],
                "key_cols": ["CategoryCode", "SeriesCode"]
            },
            "Изделия": {
                "handler": self.db_manager.add_item_number_description,
                "get_all_from_db": self.db_manager.get_item_number_descriptions,
                "columns": ["CategoryCode", "SeriesCode", "ItemNumberCode", "ItemNumberName"],
                "required_cols": ["CategoryCode", "SeriesCode", "ItemNumberCode", "ItemNumberName"],
                "key_cols": ["CategoryCode", "SeriesCode", "ItemNumberCode"]
            },
            "Операции": {
                "handler": self.db_manager.add_operation_description,
                "get_all_from_db": self.db_manager.get_operation_descriptions,
                "columns": ["OperationCode", "OperationName"],
                "required_cols": ["OperationCode", "OperationName"],
                "key_cols": ["OperationCode"]
            },
        }

    def _generate_key(self, row_data, key_cols):
        """Генерирует уникальный ключ из данных строки на основе ключевых столбцов."""
        # Ensure all key_cols are present in row_data before joining
        if not all(col in row_data for col in key_cols):
            return None  # Or raise an error, depending on desired strictness
        return "-".join(str(row_data[col]) for col in key_cols)

    def import_from_excel(self, excel_file):
        """
        Выполняет интеллектуальный импорт из Excel:
        - Добавляет новые записи в базу данных.
        - Обновляет описания существующих записей.
        - Сообщает о записях, присутствующих в БД, но отсутствующих в Excel.
        - Не удаляет существующие оснастки.
        Возвращает (success_status, added_counts, updated_counts, skipped_counts, missing_from_excel_data).
        """
        if not os.path.exists(excel_file):
            print(f"Ошибка: Файл Excel '{excel_file}' не найден.")
            return False, {}, {}, {}, {}

        try:
            workbook = openpyxl.load_workbook(excel_file)
        except Exception as e:
            print(f"Ошибка при открытии файла Excel '{excel_file}': {e}")
            return False, {}, {}, {}, {}

        print(f"\n--- Начинаем интеллектуальный импорт данных из '{excel_file}' ---")

        added_counts = {sheet_name: 0 for sheet_name in self.sheet_configs.keys()}
        updated_counts = {sheet_name: 0 for sheet_name in self.sheet_configs.keys()}
        skipped_counts = {sheet_name: 0 for sheet_name in self.sheet_configs.keys()}
        missing_from_excel_data = {sheet_name: [] for sheet_name in self.sheet_configs.keys()}

        overall_success = True

        for sheet_name, config in self.sheet_configs.items():
            print(f"Обработка листа: '{sheet_name}'")
            if sheet_name not in workbook.sheetnames:
                print(f"  Предупреждение: Лист '{sheet_name}' не найден в файле Excel. Пропуск.")
                overall_success = False
                continue

            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            header_to_col_idx = {header: idx for idx, header in enumerate(headers)}

            excel_data_for_sheet = {}  # To store data from current Excel sheet by key

            # Read data from Excel
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                is_row_valid = True
                for col_name in config["columns"]:
                    col_idx = header_to_col_idx.get(col_name)
                    if col_idx is not None:
                        cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                        row_data[col_name] = cell_value
                    else:
                        print(f"  Предупреждение: Колонка '{col_name}' не найдена на листе '{sheet_name}'.")
                        is_row_valid = False
                        break

                # Basic validation for required columns
                if not all(row_data.get(col) is not None for col in config["required_cols"]):
                    print(
                        f"  Строка {row_idx}: Пропуск из-за отсутствия данных в обязательных колонках: {config['required_cols']}.")
                    skipped_counts[sheet_name] += 1
                    is_row_valid = False

                if is_row_valid:
                    key = self._generate_key(row_data, config["key_cols"])
                    if key:
                        excel_data_for_sheet[key] = row_data
                    else:
                        print(f"  Строка {row_idx}: Не удалось сгенерировать ключ для записи. Пропуск.")
                        skipped_counts[sheet_name] += 1
                        overall_success = False

            # Fetch existing data from DB for this sheet type
            existing_db_data_for_sheet = {
                self._generate_key(d, config["key_cols"]): d
                for d in config["get_all_from_db"]()
                if self._generate_key(d, config["key_cols"]) is not None
            }

            # Compare Excel data with DB data
            for excel_key, excel_row in excel_data_for_sheet.items():
                if sheet_name == "Категории":
                    status = self.db_manager.add_category(excel_row["CategoryCode"], excel_row["CategoryName"])
                elif sheet_name == "Серии":
                    status = self.db_manager.add_series_description(excel_row["CategoryCode"], excel_row["SeriesCode"],
                                                                    excel_row["SeriesName"])
                elif sheet_name == "Изделия":
                    status = self.db_manager.add_item_number_description(excel_row["CategoryCode"],
                                                                         excel_row["SeriesCode"],
                                                                         excel_row["ItemNumberCode"],
                                                                         excel_row["ItemNumberName"])
                elif sheet_name == "Операции":
                    status = self.db_manager.add_operation_description(excel_row["OperationCode"],
                                                                       excel_row["OperationName"])
                else:
                    status = 'error'  # Should not happen with defined sheet_configs

                if status == 'added':
                    added_counts[sheet_name] += 1
                elif status == 'updated':
                    updated_counts[sheet_name] += 1
                elif status == 'skipped':
                    skipped_counts[sheet_name] += 1
                elif status == 'error':
                    overall_success = False
                    print(f"  Ошибка при обработке записи из Excel: {excel_row}")

            # Identify entries in DB but missing from Excel
            for db_key, db_row in existing_db_data_for_sheet.items():
                if db_key not in excel_data_for_sheet:
                    missing_from_excel_data[sheet_name].append(db_row)

        print("\n--- Отчет по импорту ---")
        for sheet_name in self.sheet_configs.keys():
            print(f"Лист '{sheet_name}':")
            print(f"  Добавлено записей: {added_counts[sheet_name]}")
            print(f"  Обновлено записей: {updated_counts[sheet_name]}")
            print(f"  Пропущено записей (не изменились или ошибки в Excel): {skipped_counts[sheet_name]}")
            if missing_from_excel_data[sheet_name]:
                print(f"  Записи в БД, отсутствующие в Excel:")
                for item in missing_from_excel_data[sheet_name]:
                    print(f"    {item}")

        print("--- Импорт данных завершен ---")
        return overall_success, added_counts, updated_counts, skipped_counts, missing_from_excel_data


# Example usage for testing (can be removed in final version)
if __name__ == "__main__":
    db_name = "my_fixtures_app_test.db"
    base_db_dir = "fixture_database_root_app_test"
    excel_file = "classifier_data.xlsx"  # Make sure this file exists for testing

    # Clean up previous test DB for a fresh start
    if os.path.exists(base_db_dir):
        shutil.rmtree(base_db_dir)
    os.makedirs(base_db_dir, exist_ok=True)

    db_manager = FixtureDBManager(db_name=db_name, base_db_dir=base_db_dir)
    importer = ExcelClassifierImporter(db_manager)

    print("\n--- Первый импорт (ожидаем добавление всех данных) ---")
    success, added, updated, skipped, missing = importer.import_from_excel(excel_file)
    if success:
        print("Первый импорт успешен.")
        print(f"Добавлено: {added}")
        print(f"Обновлено: {updated}")
        print(f"Пропущено: {skipped}")
        print(f"Отсутствует в Excel (после первого импорта, должно быть пусто): {missing}")
    else:
        print("Первый импорт не удался.")

    # Add some dummy data directly to DB to simulate changes not in Excel
    db_manager.add_category("ZZ", "Test Category Z")
    db_manager.add_operation_description("OPZ", "Test Operation Z")
    # Simulate a change in Excel for an existing item (e.g., change name of CS category)
    # This would require manually editing classifier_data.xlsx for testing this specific scenario.

    print("\n--- Второй импорт (ожидаем пропуск существующих, обновление измененных и отчет по отсутствующим) ---")
    success, added, updated, skipped, missing = importer.import_from_excel(excel_file)
    if success:
        print("Второй импорт успешен.")
        print(f"Добавлено во втором импорте (должно быть 0 для существующих): {added}")
        print(f"Обновлено во втором импорте (зависит от изменений в Excel): {updated}")
        print(f"Пропущено во втором импорте: {skipped}")
        print(f"Отсутствует в Excel (ожидаем ZZ и OPZ, если они не были добавлены в Excel): {missing}")
    else:
        print("Второй импорт не удался.")

    db_manager.close()
